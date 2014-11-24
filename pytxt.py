from __future__ import division
import re,os



def num2col(i):
	import string
	alpha=string.ascii_uppercase
	if i<len(alpha):
		code=alpha[i]
	else:
		import math
		offset=int(math.floor(i / len(alpha)))
		code1=alpha[offset - 1]
		offset2=i - (len(alpha) * offset)
		code2=alpha[offset2]
		code=code1+code2
	return code

def yank(text,tag,none=None):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)
	
	try:
		return text.split(tag[0])[1].split(tag[1])[0]
	except IndexError:
		return none


import re
import htmlentitydefs

def safeunicode(x):
	y=u''
	for xx in x:
		if xx=="'":
			y+="-"
			continue
		try:
			y+=xx.encode('utf-8')
		except:
			y+='-'
	return y

def list2freqs(list,tfy=False):
	x=[_x.strip() for _x in list.split('\n')]
	return toks2freq(x,tfy=tfy)

def convertentity(m):
	if m.group(1)=='#':
		try:
			return chr(int(m.group(2)))
		except ValueError:
			return '&#%s;' % m.group(2)
	try:
		return htmlentitydefs.entitydefs[m.group(2)]
	except KeyError:
		return '&%s;' % m.group(2)

def converthtml(s):
	return re.sub(r'&(#?)(.+?);',convertentity,s)

def escape(s):
	return converthtml(s).replace('\ufffd','')


def yanks(text,tag):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)

	return [ x.split(tag[1])[0] for x in text.split(tag[0])[1:] ]

def yanks2(text,tag):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)

	ys=[]
	#return [ tag[0][-1].join(x.split(tag[0][:-1])[1].split(tag[0][-1])[1:]) for x in text.split(tag[1])[:-1] ]
	
	for x in text.split(tag[1])[:-1]:
		try:
			x=x.split(tag[0][:-1])[1].split(tag[0][-1])[1:]
			x=tag[0][-1].join(x)
		except IndexError:
			pass
		ys.append(x)
	return ys

class String(unicode):
	def __str__(self):
		return self.str
	def __repr__(self):
		return self.str
	def lines(self,lb='\n'):
		return self.str.split(lb)
	def __init__(self,str,**meta):
		for k,v in meta.items(): setattr(self,k,v)
		self.str=str

class StringObject(object):
	def __str__(self):
		return self.str.encode('utf-8')
	def __unicode__(self):
		return self.str
	def __repr__(self):
		return self.str
	def lines(self,lb='\n'):
		return self.str.split(lb)
	def __init__(self,str,**meta):
		for k,v in meta.items(): setattr(self,k,v)
		self.str=str
	def split(self,x):
		return self.str.split(x)

def pull1(text,tag,returnType='u',html=True):
	x=pull(text,tag,returnType)
	if not len(x): return None
	xx=x[0]
	if not html: xx=unhtml(xx)
	return xx
	

# def tags(html):
# 	def next(tag):
# 		if 
# 		
# 	import bs4
# 	bs=bs4.BeautifulSoup(html)
	

def pull(text,tag,returnType='u'):
	#print tag,returnType
	#if not returnType: returnType='u'
	if type(tag)==type(''): tag=tagname2tagtup(tag)
	ys=[]
	
	## split by closing tag, excluding the last [b/c outside of tag]
	for x in text.split(tag[1])[:-1]:
		## now try to split by opening tag omitting closing ">"
		x=x.split(tag[0][:-1]+' ')
		
		if len(x)>1:	# then there is such a thing as "<tag "
			x=x[1]
			## now split by closing ">"
			y=x.split(tag[0][-1])
			## get attrs
			attrs=y[0]
			adict={}
			if attrs:
				key=None
				for ax in attrs.split():
					if '=' in ax:
						axs=ax.split('=')
						k=axs[0]
						v='='.join(axs[1:])
						#if not "'" in v and not '"' in v:
						adict[k]=v.replace('"','').replace("'",'')
						key=k
					elif key:
						adict[key]+=' '+ax.replace('"','').replace("'",'')
		
			for k in adict:
				adict[k]=adict[k].strip()
				if adict[k].isdigit(): adict[k]=int(adict[k])
		
			## rejoin text by closing ">", without attrs
			text=tag[0][-1].join(y[1:])
		else:
			text=x[0].split(tag[0])[1]
			adict={}

		if returnType=='u':
			string=text
		else:
			string=StringObject(text)
			for k,v in adict.items(): setattr(string,k,v)
		ys.append(string)
	return ys
	
	

def tagname2tagtup(tagname):
	return ('<'+tagname+'>','</'+tagname+'>')

def safestr(string):
	try:
		return str(string)
	except UnicodeEncodeError:
		return str(string.encode('utf-8','replace'))
	except:
		return "<????>"

def is_safe(string):
	try:
		return str(string)==ascii(string)
	except:
		return False


def simple(str):
	o=[]
	for x in str:
		try:
			unicode(x)
			o+=[x]
		except:
			pass
	return ''.join(o)

def ascii(inputstring):
	o=[]
	for x in inputstring:
		try:
			str(x)
			o+=[x]
		except:
			pass
	return ''.join(o)

def rascii(inputstring,woops='.'):
	o=[]
	for x in inputstring:
		try:
			str(x)
			o+=[x]
		except:
			o+=[woops]
	return ''.join(o)

def dict2xml(d,root="xml"):
	o=[]
	for k,v in sorted(d.items(),reverse=False):
		o+=["<"+k+">"+v+"</"+k+">"]
	return "<"+root+">\n\t"+ "\n\t".join(o) + "\n</"+root+">"
	

def neginback(strnum):
	if strnum.startswith("-"):
		return strnum[1:]+"-"
	else:
		return strnum

def thetime():
	from time import localtime, strftime
	return strftime("%Y%m%d.%H%M", localtime())

# these two lists serves as building blocks to construt any number
# just like coin denominations.
# 1000->"M", 900->"CM", 500->"D"...keep on going 
decimalDens=[1000,900,500,400,100,90,50,40,10,9,5,4,1]
romanDens=["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]


def roman(dec):
	"""
	Perform sanity check on decimal and throws exceptions when necessary
	"""		
        if dec <=0:
	  raise ValueError, "It must be a positive"
         # to avoid MMMM
	elif dec>=4000:  
	  raise ValueError, "It must be lower than MMMM(4000)"

	return decToRoman(dec,"",decimalDens,romanDens)

def decToRoman(num,s,decs,romans):
	"""
	  convert a Decimal number to Roman numeral recursively
	  num: the decimal number
	  s: the roman numerial string
	  decs: current list of decimal denomination
	  romans: current list of roman denomination
	"""
	if decs:
	  if (num < decs[0]):
	    # deal with the rest denomination
	    return decToRoman(num,s,decs[1:],romans[1:])		  
	  else:
	    # deduce this denomation till num<desc[0]
	    return decToRoman(num-decs[0],s+romans[0],decs,romans)	  
	else:
	  # we run out of denomination, we are done 
	  return s


def flatten_ld(metald,flatten_prefix='window'):
	ld2=[]
	for d in metald:
		flatten=[k for k in d.keys() if k.startswith(flatten_prefix)]
		include_with_flatten=[]
		not_to_flatten=list(set(d.keys())-set(flatten))
		for k in [kx for kx in flatten]:
			for k2 in [kx for kx in flatten if kx.startswith(k) and kx!=k]:
				flatten.remove(k2)
				include_with_flatten.append(k2)
		for k in flatten:
			d2={}
			for dk in not_to_flatten: d2[dk]=d[dk]
			d2[flatten_prefix+'Type']=noPunc(k.replace(flatten_prefix,''))
			d2[flatten_prefix+'Value']=d[k]
			for dk in include_with_flatten:
				if not dk.startswith(k): continue
				dkx=noPunc(dk.replace(k,''))
				dkx=flatten_prefix+dkx[0].upper()+dkx[1:]
				d2[dkx]=d[dk]
			ld2.append(d2)
	return ld2



def ynk(text,start,end,inout=""):
	if (not start in text) or (not end in text):
		return ""

		
	try:
		if (inout=="in" or inout==0):
			return text.split(start)[1].split(end)[0]
		elif (inout=="out" or inout==1):
			return text.split(end)[0].split(start)[-1]
		else:
			o=[]
			for x in text.split(start):
				#if x.count(">")>1:
				#	x=x[x.index(">")+1:]
				
				xx=x.split(end)[0].strip()
				if not xx: continue
				if xx.startswith("<!DOCTYPE"): continue		# NYT hack
				if xx.startswith("<NYT_"): continue 
				if xx.startswith("<script"): continue

				o.append(xx.replace("\n"," ").replace("\r"," "))
			return "\n\n".join(o)
	except:
		return ""

		
def nupos2desc(pos):
	translation={'a':'adverb/conj/prep as adverb','av':'adverb','c':'adverb/conj/prep as conj','cc':'coordinating conjunction','cr':'numeral','cs':'subordinating conjunction / that as conj','d':'determiner [that, much]','d.':'d.?','dc':'comparative determiner [less]','dg':"determiner as possessive [the latter's]",'ds':'superlative determiner','dt':'article','dx':'negative determiner as adverb','fw':'foreign word','j':'adjective','j.':'j.?','jc':'comparative adjective','jp':'proper adjective','js':'superlative adjective','n':'adjective/participle as noun','n1':'singular noun','n2':'plural noun','ng':'possessive noun','nj':'proper adjective [Roman]','np':'proper noun','or':'ordinal number','p':'adj/conj/prep as prep','pc':'adj/conj/prep as particle','pi':'indefinite pronoun [one]','pn':'personal pronoun','po':'possessive pronoun','pp':'preposition','px':'reflexive pronoun','q':'wh-word, interrogative use','r':'wh-word, relative use','sy':'symbol','uh':'interjection','vb':'to be, any tense','vd':'to do, any tense','vh':'to have, any tense','vm':'modal verb, any tense','vv':'verb','xx':'negative','zz':'unknown token'}
	
	return translation.get(pos,pos)

def nupos():
	ld=tsv2ld('/Users/ryan/inbox/python/nupos.txt')
	nd={}
	for dx in ld:
		nd[dx['tag']]=dx
	return nd
	
def bunch_ld(ld,key):
	last_val=None
	newld=[]
	newd={}
	for d in ld:
		if not last_val or d.get(key,None)!=last_val:
			if newd: newld+=[newd]
			newd=dict(d.items())
		else:
			for k in d:
				if not k in newd or newd[k]!=d[k]:
					if type(newd[k])==list:
						newd[k].append(d[k])
					else:
						newd[k]=[newd[k], d[k]]
		last_val=d.get(key,None)
	if newd: newld+=[newd]
	return newld


def xls2ld(fn,header=[],sheetname=True,keymap={}):
	import xlrd
	headerset=True if len(header) else False
	f=xlrd.open_workbook(fn)
	ld=[]
	def _boot_xls_sheet(sheet,header=[]):
		ld2=[]
		for y in range(sheet.nrows):
			if not header:
				for xi in range(sheet.ncols):
					cell=sheet.cell_value(rowx=y,colx=xi)
					header+=[cell]
				continue
			d={}
			for key in header:
				try:
					value=sheet.cell_value(rowx=y, colx=header.index(key))
					d[key]=value
					#print key,value,y,header.index(key),row[header.index(key)]
				except:
					#print "!! "+key+" not found in "+str(sheet)
					#d[key]=''
					pass
			if len(d):
				if sheetname: d['sheetname']=sheet.name
				ld2.append(d)
		return ld2


	if f.nsheets > 1:
		sheetnames=sorted(f.sheet_names())
		for sheetname in sheetnames:
			sheet=f.sheet_by_name(sheetname)
			for d in _boot_xls_sheet(sheet,header=header if headerset else []):
				ld.append(d)			
	else:
		sheet = f.sheet_by_index(0)
		ld.extend(_boot_xls_sheet(sheet,header=header if headerset else []))

	return ld
	
def xls2dld(fn,header=[]):
	return ld2dld(xls2ld(fn,header=header,sheetname=True), 'sheetname')

def levenshtein(s1, s2):
	l1 = len(s1)
	l2 = len(s2)

	matrix = [range(l1 + 1)] * (l2 + 1)
	for zz in range(l2 + 1):
		matrix[zz] = range(zz,zz + l1 + 1)
	for zz in range(0,l2):
		for sz in range(0,l1):
			if s1[sz] == s2[zz]:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz])
			else:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz] + 1)
	return matrix[l2][l1]

def xlsx2ld(fn,header=[],numsheets=1):
	from openpyxl.reader.excel import load_workbook
	header_set=bool(len(header))
	wb=load_workbook(filename=fn)
	ld=[]
	for sheet in wb.worksheets[:numsheets]:
		if not header_set: header=[]
		#header=[]
		for rownum,row in enumerate(sheet.rows):
			values=[]
			for cell in row:
				value=cell.value
				if value is None:
					value=''
				
				try:
					value=float(value)/0
				except:
					value=value
					if not isinstance(value, unicode):
						value=unicode(value)
				values.append(value)
			if not rownum and not len(header):
				header=values
			else:
				d=dict((header[i],values[i]) for i in range(len(values)))
				ld+=[d]
	return ld
	
def dl2ld(dl,kcol='group'):
	ld=[]
	for k in dl:
		for d in dl[k]:
			d[kcol]=k
			ld+=[d]
	return ld

def fn2ld(fn,header=[],sep='\t',nsep='\n'):
	import codecs
	f=codecs.open(fn,encoding='utf-8')
	for line in f:
		line=line.strip()
		if not header:
			header=line.split(sep)
			continue
		dx={}
		for i,val in enumerate(line.split(sep)):
			key=header[i] if len(header)>i else 'key_'+str(i)
			dx[key]=val
		yield dx

def goog2tsv(googsrc):
	import bs4
	dom=bs4.BeautifulSoup(googsrc)
	header=[th.text for th in dom('thead')[0]('th')]
	header=header if True in [bool(hx) for hx in header] else None
	old=[]
	for row in dom('tbody')[0]('tr'):
		rowdat=[cell.text for cell in row('td')]
		if not header:
			header=rowdat
			#print ">> HEADER:",header
			continue
		odx=dict(zip(header,rowdat))
		old+=[odx]
	return old


def tsv2ld(fn,tsep='\t',nsep='\n',u=True,header=[],keymap={},zero='',removeEmpties=False):
	import os
	if fn.startswith('http'):
		import urllib
		f=urllib.urlopen(fn)
		t=f.read()
		if fn.endswith('/pubhtml'):
			return goog2tsv(t)
		f.close()
	elif not os.path.exists(fn):
		t=fn
	elif u:
		import codecs
		f=codecs.open(fn,encoding='utf-8')
		t=f.read()
		f.close()
	else:
		f=open(fn,'r')
		t=f.read()
		f.close()
	t=t.replace('\r\n','\n')
	t=t.replace('\r','\n')
	
	#header=[]
	listdict=[]
	
	
	for line in t.split(nsep):
		if not line.strip(): continue
		line=line.replace('\n','')
		ln=line.split(tsep)
		#print ln
		if not header:
			header=ln
			for i,v in enumerate(header):
				if v.startswith('"') and v.endswith('"'):
					header[i]=v[1:-1]
			continue
		edict={}
		for i in range(len(ln)):
			try:
				k=header[i]
			except IndexError:
				#print "!! unknown column for i={0} and val={1}".format(i,ln[i])
				continue
			v=ln[i].strip()
			
			if k in keymap:
				print v, type(v)
				v=keymap[k](v)
				print v, type(v)
			else:
				if v.startswith('"') and v.endswith('"'):
					v=v[1:-1]
				try:
					v=float(v)
				except ValueError:
					v=v
			
			if type(v) in [str,unicode] and not v:
				if zero=='' and removeEmpties:
					continue
				else:
					v=zero
			edict[k]=v
		if edict:
			listdict.append(edict)
	return listdict
	
def dkey(d,extra={}):
	#import pytxt
	#kv='__'.join(['{0}_{1}'.format(unicode(k),unicode(v)) for k,v in sorted(d.items()+extra.items())])
	import cPickle
	kv=cPickle.dumps(dict(d.items()+extra.items()))
	return kv
	
def unhtml(data):
	if not data: return data
	try:
		from lxml.html import fromstring
		return fromstring(data).text_content()
	except:
		return remove_html_tags(data)

def remove_html_tags(data):
	#data=safestr(data)
	p=re.compile(r'<.*?>',re.UNICODE)
	try:
		y=str(p.sub('',data)).strip().split('">')
	except UnicodeEncodeError:
		y=unicode(p.sub('',data)).strip().split('">')
	while(('&' in y) and (';' in y)):
		y=y[:y.index('&')]+y[y.index(';')+1:]
	try:
		return y[1].strip()
	except:
		return y[0]

def htm2txt(element):
    import types
    text = ''
    for elem in element.recursiveChildGenerator():
        if isinstance(elem, types.StringTypes):
            txt=elem.strip().replace('\r\n',' ').replace('\r',' ').replace('\n',' ')
            while '  ' in txt: txt=txt.replace('  ',' ')
            text += txt
        elif elem.name in ['br','p']:
            text += '\n'
    return text

def extractTags(text,leavetexttags=[u"placeName"]):
	tags=[]
	tags_milestone=[]
	yankeds=[]
	
	if "</" in text:
		for x in text.split("</")[1:]:
			tags.append(x.split(">")[0])
	
	if "/>" in text:
		for x in text.split("/>")[:-1]:
			x=x.split("<")[-1]
			try:
				x=x.split()[0]
			except IndexError:
				x=x
			#if "/" in x: continue
			#if not x: continue
			tags_milestone.append(x)

	for tag in tags_milestone:
		yanked=yank(text,("<"+tag,"/>"))
		while yanked.strip():
			ydat="<"+tag+yanked+"/>"
			#yankeds.append(ydat)
			text=text.replace(ydat,' ')
			yanked=yank(text,("<"+tag,"/>"))
	
	for tag in tags:
		yanked=yank(text,("<"+tag,"</"+tag+">"))
		while yanked and yanked.strip():				
			ydat="<"+tag+yanked+"</"+tag+">"
			if tag in leavetexttags:
				text=text.replace(ydat,remove_html_tags(yanked.split(">")[-1]))
			else:
				yankeds.append(ydat)
				text=text.replace(ydat,' ')
			yanked=yank(text,("<"+tag,"</"+tag+">"))
	return (text.replace("\n","").replace("\r",""),yankeds)

def gleanPunc2(aToken):
	aPunct0 = ''
	aPunct1 = ''
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct0 = aPunct0+aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct1 = aToken[-1]+aPunct1
		aToken = aToken[:-1]
			
	return (aPunct0, aToken, aPunct1)



def gleanPunc(aToken):
	aPunct = None
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct = aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct = aToken[-1]
		aToken = aToken[:-1]
	return (aToken, aPunct)

def isPunc(token):
	return (len(noPunc(token))==0)

def alphas(token):
	return ''.join([x for x in token if x.isalpha() or x in [' ']])

def noPunc(token):
	if not token: return token
	x=gleanPunc(token)[0]
	x=x.split('&')[0]
	y=x.split(';')
	try:
		x=y[1]
	except IndexError:
		pass
	x=x.split('\\')[0]
	return x

def bigrams(l):
	return ngram(l,2)

def ngram(l,n=3):
	grams=[]
	gram=[]
	for x in l:
		gram.append(x)
		if len(gram)<n: continue
		g=tuple(gram)
		grams.append(g)
		gram.reverse()
		gram.pop()
		gram.reverse()
	return grams
	

def fn2dict(fn,sep='\t'):
	return readDict(fn,sep=sep)

def readDict(fn,sep='\t'):
	try:
		d={}
		#f=open(fn)
		import codecs
		f=codecs.open(fn,encoding='utf-8')
		for line in f:
			ln=line.split(sep)
			k=ln[0].strip()
			v=ln[1].strip()
			
			if v.isdigit():
				d[k]=int(v)
			else:
				d[k]=v
		
		if len(d):
			return d
		else:
			return None
		
	except IOError:
		return {}
	
def writeDict(fn,d,sep="\t",toprint=True):
	o=""
	for k,v in d.items():
		o+=sep.join(str(x) for x in [k,v])+"\n"
	write(fn,o,toprint)


def dict2tuple(d):
	return tuple((k,v if type(v)!=list else tuple(v)) for k,v in sorted(d.items()))


def extractTagsAsDict(text,leavetexttags=[u"placeName"]):
	text,tags=extractTags(text,leavetexttags)
	tagdict={}
	for tag in tags:
		
		opentag=tag.split(">")[0].split("<")[1].strip()
		tagbody=unhtml(tag).strip()
		
		if not tagbody: continue
		
		if " " in opentag:
			spaces=opentag.split()
			tagname=spaces[0]
			for space in spaces[1:2]:
				if not space.strip(): continue
				dat=space.strip().split("=")
				k=dat[0]
				try:
					v=dat[1]
				except:
					print "error with extractTagsAsDict in pytxt"
					continue
				v=v.replace('"','').replace("'","").strip()
				
				try:
					tagdict[tagname][k][v]=tagbody
				except KeyError:
					try:
						tagdict[tagname][k]={}
						tagdict[tagname][k][v]=tagbody
					except KeyError:
						tagdict[tagname]={}
						tagdict[tagname][k]={}
						tagdict[tagname][k][v]=tagbody
		
		else:
			tagname=opentag
			tagdict[tagname]=tagbody	
				
	return tagdict


def writeToFile(folder,fn,data,extension="tsv"):
	#ofolder=os.path.join(folder,'results','stats','corpora',name)

	if not os.path.exists(folder):
		os.makedirs(folder)

	ofn=os.path.join(folder,'.'.join([fn,extension]))
	print ">> saved:",ofn
	of = open(ofn,'w')
	of.write(data)
	of.close()


def chunk(fn,num=10):
	import codecs,os
	t=codecs.open(fn,encoding='utf-8').read()
	words=t.split()
	numwords=len(words)
	wordperseg=int(numwords/num)
	for segnum in range(num):
		segwords=words[segnum*wordperseg : (segnum+1)*wordperseg]
		segtext=' '.join(segwords)
		segfn=os.path.basename(fn).replace('fulltext',str(segnum+1))
		segfn=segfn.split('_')[0] + '_' + segfn.split('_')[1].zfill(2) + '_' + '_'.join(segfn.split('_')[2:])
		write(segfn,segtext)


def split_texts(infolder,outfolder,lim=1000):
	for fn in os.listdir(infolder):
		import codecs
		text=codecs.open(os.path.join(infolder,fn),encoding='utf-8').read().split()
		for n,txt in enumerate(segment(text,lim)):
			ofn=fn.replace('.txt','.'+str(n).zfill(4)+'.txt')
			write(os.path.join(outfolder,ofn), ' '.join(txt))
		
	

def segment(l,num=200):
	import math
	segments_needed=int(math.ceil(len(l)/num))
	for n in range(segments_needed+1):
		yield l[n*num:(n+1)*num]

def dld2dll(dld):
	for k in dld:
		dld[k]=ld2ll(dld[k])
	return dld

def write_xls(fn,data,sheetname='index',toprint=True,limFields=None,widths=[],zero=''):
	import xlwt
	wb=xlwt.Workbook(encoding='utf-8')
	
	if datatype(data).startswith('ld'):
		dd={}
		dd[sheetname]=ld2ll(data,zero=zero)
	elif type(data)!=type({}):
		dd={}
		dd[sheetname]=data
	elif datatype(data).startswith('dld'):
		dd=dld2dll(data)
	else:
		dd=data
	
	for sheetname,data in sorted(dd.items()):
		ws=wb.add_sheet(sheetname)
		nr=-1
		#style = xlwt.easyxf('align: wrap True')
		#style=xlwt.easyxf('')
		for row in data:
			nc=-1
			nr+=1
			for cell in row:
				nc+=1
				if not (type(cell)==type(1) or type(cell)==type(1.0)):
					try:
						ws.row(nr).set_cell_text(nc,cell)
					except TypeError:
						ws.row(nr).set_cell_text(nc,unicode(cell))
				else:
					ws.row(nr).set_cell_number(nc,cell)	
	wb.save(fn)
	if toprint:
		print ">> saved:",fn
	

def tmp(data):
	import tempfile
	f=tempfile.NamedTemporaryFile()
	f.write(data)
	#f.close()
	return f

def write_tmp(data,suffix=''):
	import time
	fn='/Lab/Processing/tmp/'+str(time.time()).replace('.','')+suffix
	write(fn,data)
	return fn	

def ld2html(ld):
	keys=ld2keys(ld)
	headerrow=['<th>%s</th>'%k for k in keys]
	rows=[]
	rows+=['\n\t\t'.join(headerrow)]
	for d in ld:
		row=['<td>%s</td>'%d.get(k,'') for k in keys]
		rows+=['\n\t\t'.join(row)]
	ostr=u"<table>\n\t<tr>\n\t\t" + u'\n\t</tr>\n\t<tr>\n\t\t'.join(rows) + u"\n\t</tr>\n</table>"
	return ostr

def ld2keys(ld):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=list(sorted(list(set(keys))))
	return keys

def ld2ll(ld,zero='',tostr=False,uni=True):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=sorted(list(set(keys)))
	o=[keys]
	for d in ld:
		l=[]
		for k in keys:
			v=d.get(k,zero)
			if tostr: 
				v=unicode(v) if uni else str(v)
			l+=[v]
		o+=[l]
	return o
	

def write_ld(fn,ld,zero='',timestamp=None):
	return write(fn,ld2ll(ld,zero=zero),timestamp=timestamp)

def dd2ld(dd,rownamecol='rownamecol'):
	if not rownamecol:
		return [ (dict(v.items())) for k,v in dd.items() ]
	else:
		return [ (dict(v.items() + [(rownamecol,k)])) for k,v in dd.items() ]

def dld2ld(dld,key='rownamecol'):
	ld=[]
	for k in dld:
		for d in dld[k]:
			d[key]=k
			ld+=[d]
	return ld

def ld_resample(ld,key='rownamecol',n=None):
	import random
	dld=ld2dld(ld,key)
	minlen_actually=min([len(dld[k]) for k in dld])
	minlen=minlen_actually if not n or n>minlen_actually else n
	ld2=[]
	for k in dld:
		ld2+=random.sample(dld[k],minlen)
	return ld2

def ld2dld(ld,key='rownamecol'):
	dld={}
	for d in ld:
		if not d[key] in dld: dld[d[key]]=[]
		dld[d[key]]+=[d]
	return dld

def ld2dd(ld,rownamecol='rownamecol'):
	dd={}
	for d in ld:
		dd[d[rownamecol]]=d
		#del dd[d[rownamecol]][rownamecol]
	return dd

def datatype(data,depth=0,v=False):
	def echo(dt):
		if not v: return
		for n in range(depth): print "\t",
		print '['+dt[0]+']'+dt[1:],
		try:
			print "[{0} records]".format(len(data),dt)
		except:
			print

	if type(data) in [str,unicode]:
		echo('string')		
		return 's'
	elif type(data) in [float,int]:
		echo('number')
		return 'n'
	elif type(data) in [list]:
		echo('list')
		if not len(data):
			return 'l'
		else:
			return 'l'+datatype(data[0],depth=depth+1,v=v)
	elif type(data) in [dict]:
		echo('dictionary')
		if not len(data):
			return 'd'
		else:
			return 'd'+datatype(data.values()[0],depth=depth+1,v=v)
	else:
		print "WHAT TYPE OF DATA IS THIS:"
		print data
		print type(data)
		print
		return '?'


def limcols(ld,limcol=255):
	keyd={}
	keys=set()
	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-keys:
			keyd[key]=0
		keys|=dkeys
		for k in d:
			if d[k]:
				keyd[k]+=1

	cols=set(sorted(keyd.keys(), key=lambda _k: (-keyd[_k],_k))[:limcol])

	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-cols:
			del d[key]

	return ld

def ld2str(ld,**data):
	if data['limcol']:
		print ">> limiting columns"
		limcol=data['limcol']
		ld=limcols(ld,limcol)
	if 'limcol' in data:
		del data['limcol']
	return ll2str(ld2ll(ld),**data)

def d2ll(d):
	try:
		return [[k,v] for k,v in sorted(d.items(),key=lambda lt: -lt[1])]
	except:
		return [[k,v] for k,v in d.items()]

def d2str(d,uni=True):
	return ll2str(d2ll(d),uni=uni)

def strmake(x,uni=True):
	if uni and type(x) in [unicode]:
		return x
	elif uni and type(x) in [str]:
		return x.decode('utf-8',errors='replace')
	elif uni:
		return unicode(x)
	elif not uni and type(x) in [str]:
		return x
	elif not uni and type(x) in [unicode]:
		return x.encode('utf-8',errors='replace')
	
	print [x],type(x)
	return str(x)


def ll2str(ll,uni=True,join_line=u'\n',join_cell=u'\t'):
	if not uni:
		join_line=str(join_line)
		join_cell=str(join_cell)
		quotechar='"' if join_cell==',' else ''
	else:
		quotechar=u'"' if join_cell==',' else u''

	for line in ll:
		lreturn=join_cell.join([quotechar+strmake(cell,uni=uni)+quotechar for cell in line])+join_line
		yield lreturn

def l2str(l,uni=True,join_line=u'\n',join_cell=u'\t',quotechar=''):
	for line in l: yield strmake(line)+join_line

def write_ld2(fn,gen1,gen2,uni=True,badkeys=[]):
	def find_keys(gen):
		keys=set()
		for d in gen:
			keys=keys|set(d.keys())
		keys=keys-set(badkeys)
		return keys

	keys=list(sorted(list(find_keys(gen1))))
	numk=len(keys)
	
	import codecs
	of=codecs.open(fn,'w',encoding='utf-8')
	of.write('\t'.join([strmake(x) for x in keys]) + '\n')

	for d in gen2:
		data=[d.get(key,'') for key in keys]
		of.write('\t'.join([strmake(x) for x in data]) + '\n')
	of.close()
	print ">> saved:",fn


def write2(fn,data,uni=True,join_cell=u'\t',join_line=u'\n',limcol=None):
	## pass off to other write functions if necessary
	if fn.endswith('.xls'): return write_xls(fn,data)
	if fn.endswith('.csv'): join_cell=','
	
	## get datatyoe
	dt=datatype(data)
	
	## get str output for datatype
	if dt.startswith('ld'):
		o=ld2str(data,join_cell=join_cell,limcol=limcol)
	elif dt.startswith('dl'):
		o=dl2str(data,uni=uni)
	elif dt.startswith('ll'):
		o=ll2str(data,uni=uni)
	elif dt.startswith('dd'):
		o=dd2str(data,uni=uni)
	elif dt.startswith('l'):
		o=l2str(data,uni=uni)
	elif dt.startswith('d'):
		o=d2str(data,uni=uni)
	else:
		o=data
	
	## write
	import codecs
	of = codecs.open(fn,'w',encoding='utf-8') if True else open(fn,'w')
	for line in o: of.write(line)
	of.close()
	print '>> saved:',fn


def now(now=None):
	import datetime as dt
	if not now:
		now=dt.datetime.now()
	elif type(now) in [int,float,str]:
		now=dt.datetime.fromtimestamp(now)

	return '{0}{1}{2}-{3}{4}.{5}'.format(now.year,str(now.month).zfill(2),str(now.day).zfill(2),str(now.hour).zfill(2),str(now.minute).zfill(2),str(now.second).zfill(2))

def striphtml(data):
	import re
	p = re.compile(r'<.*?>')
	return p.sub('', data)




def write(fn,data,uni=True,toprint=True,join_line='\n',join_cell='\t',timestamp=None):
	if timestamp:
		from datetime import datetime
		ts=datetime.now().strftime('%Y-%m-%d_%H%M')
		fn='.'.join(fn.split('.')[:-1]) + '.' + ts + '.' + fn.split('.')[-1]

	if not uni:
		of = open(fn,'w')
	else:
		join_line=u'\n'
		join_cell=u'\t'
		import codecs
		of = codecs.open(fn,'w',encoding='utf-8')
	
	if type(data) in [list,tuple]:
		o=""
		for x in data:
			if type(x) in [list,tuple]:
				z=[]
				for y in x:
					if not uni and type(y)==type(u''):
						y=y.encode('utf-8')
					z+=[y]
				x=z
				
				try:
					line=join_cell.join(x)
				except TypeError:
					line=[]
					for y in x:
						try:
							yx=y.decode('utf-8')
						except:
							try:
								yx=str(y)
							except:
								yx=y
						line+=[yx]
					line=join_cell.join(line)
					# 
					# if not uni:
					# 	line=join_cell.join(str(y) for y in x)
					# else:
					# 	line=join_cell.join(unicode(y) for y in x)
			else:
				try:
					line=str(x)
				except UnicodeEncodeError:
					line=x.encode('utf-8')
			line=line.replace('\r','').replace('\n','')
			#print [line+join_line], type(line)
			#print [join_line], type(join_line)
			of.write(line + join_line)
	else:
		try:
			o=str(data)
		except UnicodeEncodeError:
			o=unicode(data)
		of.write(o)
	of.close()
	if toprint:
		print ">> saved:",fn.encode('utf-8')
	
def uwrite(fn,data,toprint=True,join_line='\n',join_cell='\t'):
	if type(data)==type([]):
		o=u""
		for x in data:
			if type(x)==type([]):
				z=[]
				for y in x:
					if type(y)!=type(u''):
						try:
							y=y.decode('utf-8')
						except AttributeError:
							y=unicode(y)
					z+=[y]
				x=z
				line=join_cell.join(x)
				
			else:
				if type(x)!=type(u''):
					try:
						line=x.decode('utf-8')
					except AttributeError:
						line=unicode(x)
				else:
					line=x
			line=line.replace('\n','\r').replace('\n','')
			o+=line+join_line
	else:
		o=unicode(data)

	import codecs
	of = codecs.open(fn,'w',encoding='utf-8')
	of.write(o)
	of.close()
	if toprint: print ">> saved:",fn
	
	

def makeminlength(string,numspaces):
	if len(string) < numspaces:
		for i in range(len(string),numspaces):
			string += " "
	return string
	
def get_class( kls ):
    parts = kls.split('.')
    module = ".".join(parts[:-1])
    m = __import__( module )
    for comp in parts[1:]:
        m = getattr(m, comp)            
    return m

def gleanPunc(aToken):
	aPunct = None
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct = aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct = aToken[-1]
		aToken = aToken[:-1]
	return (aToken, aPunct)
	

def count(string, look_for):
    start   = 0
    matches = 0

    while True:
        start = string.find (look_for, start)
        if start < 0:
            break

        start   += 1
        matches += 1

    return matches

def slice(l,num_slices=None,slice_length=None,runts=True,random=False):
	"""
	Returns a new list of n evenly-sized segments of the original list
	"""
	if random:
		import random
		random.shuffle(l)
	if not num_slices and not slice_length: return l
	if not slice_length: slice_length=int(len(l)/num_slices)
	newlist=[l[i:i+slice_length] for i in range(0, len(l), slice_length)]
	if runts: return newlist
	return [lx for lx in newlist if len(lx)==slice_length]


def select2(dict_with_key_as_option_and_value_as_prob):
	import random
	import bisect
	import collections

	def cdf(weights):
	    total=sum(weights)
	    result=[]
	    cumsum=0
	    for w in weights:
	        cumsum+=w
	        result.append(cumsum/total)
	    return result

	def choice(population,weights):
	    assert len(population) == len(weights)
	    cdf_vals=cdf(weights)
	    x=random.random()
	    idx=bisect.bisect(cdf_vals,x)
	    return population[idx]

	items=dict_with_key_as_option_and_value_as_prob.items()
	weights=[v for k,v in items]
	population=[k for k,v in items]
	return choice(population,weights)

def select(dict_with_key_as_option_and_value_as_prob):
	import random
	d=dict_with_key_as_option_and_value_as_prob
	r = random.uniform(0, sum(d.itervalues()))
	s = 0.0
	for k, w in d.iteritems():
		s += w
		if r < s: return k
	return k


def choose(optionlist,msg="please select from above options [using commas for individual selections and a hyphen for ranges]:\n"):
	seldict={}

	selnum=0
	print
	print

	if type(optionlist)==type([]):
		for option in optionlist:
			selnum+=1
			seldict[selnum]=option
			print "\t"+"\t".join(str(x) for x in [selnum,option])
	elif type(optionlist)==type({}):
		for option,desc in optionlist.items():
			selnum+=1
			seldict[selnum]=option
			print "\t"+"\t".join(str(x) for x in [selnum,option,desc])

	inp=raw_input("\n\t>> "+msg+"\n\t").strip()
	sels=[]
	for np in inp.split(","):
		np=np.strip()
		if "-" in np:
			try:
				nn=np.split("-")
				for n in range(int(nn[0]),int(nn[1])+1):
					sels.append(seldict[n])
			except:
				continue
		else:
			try:
				sels.append(seldict[int(np)])
			except:
				continue

	return sels

def toks2str(tlist,uni=False):
	toks=[]
	putleft=False
	#print tlist
	for tk in tlist:
		tk=tk.strip()
		if not tk: continue
		tk = tk.split()[-1]
		if not tk: continue
		if (not len(toks)):
			toks+=[tk]
		elif putleft:
			toks[-1]+=tk
			putleft=False
		elif tk=='`':
			toks+=[tk]
			putleft=True
		elif tk=='-LRB-':
			toks+=['(']
			putleft=True
		elif tk=='-RRB-':
			toks[-1]+=')'
		elif len(tk)>1 and tk[-2]=="'":
			toks[-1]+=tk
		elif tk[0].isalnum():
			toks+=[tk]
		elif tk.startswith('<') and '>' in tk:
			toks+=[tk]
		else:
			toks[-1]+=tk
	if uni: return u' '.join(toks)
	return ' '.join(toks)

def escape_punc(punc,keep=['_']):
	if len(punc)>1:
		newword=[]
		for li,letter in enumerate(punc):
			newletter=escape_punc(letter)
			newword+=[newletter]
			if newletter!=letter and li+1!=len(punc):
				newword+=['_']
		return ''.join(newword)
	if punc in keep: return punc
	puncd={"'":"p_apos", ",":"p_comma", "!":"p_exclam", "-":"p_hyphen", ".":"p_period", "?":"p_ques", '"':"p_quote", ";":"p_semi"}
	puncd['_']='p_underscore'
	puncd[u'\u2014']='p_emdash'
	puncd['|']='p_pipe'
	puncd[':']='p_colon'
	puncd['&']='p_ampersand'
	puncd[u'\u201d']='p_rightquote'
	puncd[u'\u201c']='p_leftquote'
	puncd[']']='p_rightbracket'
	puncd['[']='p_leftbracket'
	puncd[')']='p_rightparenthesis'
	puncd['(']='p_leftparenthesis'
	puncd[u'\u2018']='p_leftquote_single'
	puncd[u'\u2019']='p_rightquote_single'
	puncd['`']='p_tilde'
	puncd['$']='p_dollarsign'
	if punc.isdigit():
		return 'num_'+str(punc)
	return puncd.get(punc,punc)

def mdwgo(folder='.'):
	for fn in os.listdir(folder):
		if not 'fmt_long' in fn: continue
		if not fn.endswith('.txt'): continue
		#mdw2viz(fn,cellcol='p-value_indiv',classes=True,minobs=30)
		mdw2viz(fn,cellcol='obs/exp',classes=True,minobs=30)
		

def mdw2net2(fn,minobs=3):
	import networkx as nx
	if '*' in fn:
		a,b=fn.split('*',1)
		fndir=os.path.split(fn)[0]
		if not fndir: fndir='.'
		G=nx.DiGraph()
		for _fn in os.listdir(fndir):
			if _fn.startswith(a) and _fn.endswith(b):
				print _fn
				_g=mdw2net2(os.path.join(fndir,_fn),minobs=minobs)
				for node in _g.nodes():
					if not G.has_node(node):
						G.add_node(node,**_g.node[node])
				
				for a,b,d in _g.edges(data=True): G.add_edge(a,b,**d)
		
		nx.write_gexf(G, fn+'.gexf')
		return G
	
	print ">> converting:",fn
	ld=tsv2ld(fn)
	g=nx.DiGraph()
	
	if 'Gender-Gender' in fn and 'document' in ld[0]:
		ld=[d for d in ld if d['document'] in ['Male-Male','Male-Female','Female-Female','Female-Male']]
		
	for d in ld:
		if d['obs']<minobs: continue
		for k in ['document','word']:
			nodeType=k
			node=d[k]
			if not g.has_node(node): g.add_node(node,nodeType=nodeType)
			d['weight']=1-d['p-value_indiv']
			d['weightstr']=str(round(d['weight']*100,1))+'%'
			g.add_edge(d['document'], d['word'], **d)
	nx.write_gexf(g, fn+'.minobs={0}.gexf'.format(minobs))
	return g

def mdw2net(fn,minp=0.1):
	import networkx as nx
	g=nx.Graph()
	metaphors='plague beast whore black feed laugh poison weep gold lips fiend madness'.split()
	for d in tsv2ld(fn):
		#if minp and d['p-value']>minp: continue
		for k in ['document','word']:
			nodeType=k if not d[k] in metaphors else 'word_metaphor'
			node=d[k]
			if not g.has_node(node): g.add_node(node,nodeType=nodeType)
		if d['obs/exp']>1:
			g.add_edge(d['document'], d['word'], **d)
	nx.write_gexf(g, fn+'.gexf')


def dd2mdw(fn,dd,sample=True,save=True,minnum=None):
	if sample:
		import random
		minlen=min([sum(dd[k].values()) for k in dd])
		minlen=minnum if minnum and minnum<minlen else minlen

		dl={}
		dd2={}
		for k in dd:
			words=[]
			for word,freq in dd[k].items():
				for n in range(freq): words+=[word]
			words=random.sample(words,minlen)
			dd2[k]=toks2freq(words)
	
		ld=dd2ld(dd2)
	else:
		print ">> dd2ld..."
		ld=dd2ld(dd)

	return ld2mdw(fn,ld,save=save)

def dl2mdw(fn,dl,sample=True):
	if sample:
		import random
		minlen=min([len(dl[k]) for k in dl])
		for k in dl:
			dl[k]=random.sample(dl[k],minlen)
	
	print ">> MDW:"
	for group in dl:
		print "\t>>",group,"-->",len(dl[group]),'words'
		dl[group]=toks2freq(dl[group])
	ld=dd2ld(dl)
	return ld2mdw(fn,ld)


def dd2mdw2(fn,dd,sample=True,save=False,minnum=None,maxnum=1000000):
	print ">>","dd2mdw2"
	if sample:
		import random
		#minlen=min([sum(dd[k].values()) for k in dd])
		#minlen=minnum if minnum and minnum<minlen else minlen
		#if minlen>maxnum: minlen=maxnum
		minlen=maxnum
		from scipy import stats
		

		for k in dd.keys():
			items=dd[k].items()
			weights=[x[1] for x in items]
			sumweight=float(sum(weights))
			weights=[w/sumweight for w in weights]
			population=[x[0] for x in items]
			pop_ids=[i for i,x in enumerate(population)]
			newd={}
			custm=stats.rv_discrete(name='custm', values=(pop_ids, weights))
			choices=custm.rvs(size=minlen)
			for i,x in enumerate(choices):
				xk=population[x]
				if not xk in newd: newd[xk]=0
				newd[xk]+=1

			dd[k]=newd
			print "RESAMPLED",k,len(dd[k]),sum(dd[k].values())
	
	ld=dd2ld(dd)
	print ld[0]
	return ld2mdw(fn,ld,save=save)



def ld2mdw(fn,ld,save=True):
	import rpyd2,rpy2
	print ">> ld2mdw init..."
	rekey={}
	for d in ld:
		for k in d.keys():
			try:
				if type(k)==str and rpyd2.rfy(k)==k: continue
			except rpy2.rinterface.RRuntimeError:
				pass
			#"""
			strk = k.encode('utf-8') if type(k) in [unicode] else str(k)
			hashk = 'X'+hash(strk)
			rekey[hashk]=k
			d[hashk]=d[k]
			del d[k]
	print ">> making rpyd2..."
	#print ld[0]
	r=rpyd2.RpyD2(ld,rownamecol='rownamecol',allcols=True)
	print ">> doing mdw..."
	"""
	old1,old2=r.mdw(returnType='both')
	for d in old1: d['word']=rekey.get(d['word'],d['word'])
	for d in old2: d['word']=rekey.get(d['word'],d['word'])
	if save:
		print ">> writing..."
		write2(fn.replace('.txt','.fmt_long.txt'),old1)
		write2(fn.replace('.txt','.fmt_wide.txt'),old2)
	return old1
	"""
	for d in r.mdw():
		d['word']=rekey.get(d['word'],d['word'])
		yield d

def mdw_long2wide(ld,rowcol='word',colcol='document',cellcol='obs/exp',pmin=None,minobs=3,empties=True):
	import pytxt
	dld=ld2dld(ld, rowcol)
	old=[]
	for rowname,ld in dld.items():
		if pmin and True in [d['p-value']>pmin for d in ld]: continue
		if minobs and True in [d['obs_total']<minobs for d in ld]: continue
		if not empties and True in [not d['obs_min'] for d in ld]: continue
		od={'word':rowname}
		for d in ld:
			colname=d[colcol]
			od[colname]=d[cellcol]
		old+=[od]
	
	return old

def mdw2wordnet(fn,minp=0.01,weight='p-value_indiv'):
	import networkx as nx,pytxt,pystats
	ld=fn if datatype(fn)=='ld' else tsv2ld(fn)
	g=nx.Graph()
	g2=nx.Graph()
	for i,d in enumerate(ld):
		if minp and d['p-value_indiv']>minp: continue
		print i
		g.add_edge(d['word'],d['document'],weight=d['p-value_indiv'])
	print ">> number of nodes, edges:",g.order(),g.size()
	print ">> computing shortest paths..."
	paths=nx.shortest_path(g,weight=weight)
	print ">> done."
	newedges={}
	for a in paths:
		for b in paths[a]:
			path=paths[a][b]
			print a, b, path
			if len(path)<2: continue
			pathweight=0
			for _a,_b in pytxt.bigrams(path):
				pathweight+=g[_a][_b]['weight']
			newedges[(a,b)]=pathweight
	
	zweight=pystats.zfy(newedges)
	for (a,b),weight in newedges.items():
		zw=zweight[(a,b)]
		if zw<0: continue
		g2.add_edge(a,b,weight=zw,sumweight=newedges[(a,b)])
	
	nx.write_gexf(g2, fn+'.wordnet.gexf')
	
def stanford2toks(fn,not_pos=['DT','CC','IN'],not_words=[],punc=False,lemmatize=True,replacements={},digits=False):
	import codecs,bs4
	t=codecs.open(fn,encoding='utf-8').read() if fn.endswith('.xml') else fn
	t=t.lower()
	bs=bs4.BeautifulSoup(t)
	words=[]
	for tok in bs.find_all('token'):
		pos=tok.find('pos').text.upper()
		if not_pos:
			if pos in not_pos: continue
		
		if not punc:
			if not pos[0].isalpha(): continue
		
		word=tok.find('word').text if not lemmatize else tok.find('lemma').text
		word=replacements.get(word,word)
		if not digits and word.isdigit(): continue
		if not_words and word in not_words: continue

		if not word[0].isalpha() and words:
			words[-1]+=word
		else:
			words+=[word]
	return words
	
	

def mdwlong2nozero(infolder,outfolder):
	for fn in os.listdir(infolder):
		ld=tsv2ld(os.path.join(infolder,fn))
		ld=[d for d in ld if d['obs'] and d['obs']!=0]
		write2(os.path.join(outfolder,fn), ld)
		

def mdwviz(folder='.',suffix='.fmt_long.txt',classes=True,minobs=3):
	import os
	for fn in sorted(os.listdir(folder)):
		if not fn.endswith(suffix): continue
		print fn,'...'
		#if fn<'Shakespeare.com.1598.much_ado': continue
		if os.path.exists(fn.replace(suffix,'.obsexp.minobs=5.words.pca.pdf')) and (not classes or os.path.exists(fn.replace(suffix,'.obsexp.minobs=5.classes.pvclust.pdf'))): continue
		try:
			mdw2viz(os.path.join(folder,fn),classes=classes,minobs=minobs)
		except Exception as e:
			print "!!"*50
			print "!!",e
			print "!!"*50


def mdw2viz(fn,classes=False,log=True,maxscore=5.0,cellcol='obs/exp',minobs=100):
	import rpyd2,math
	ld=tsv2ld(fn)
	if 'Gender-Gender' in fn and 'document' in ld[0]:
		print len(ld)
		ld=[d for d in ld if d['document'] in ['Male-Male','Male-Female','Female-Female','Female-Male']]
		print len(ld)
		print set(d['document'] for d in ld)
		print

	if 'document' in ld[0]: ld=mdw_long2wide(ld,cellcol=cellcol,minobs=minobs)
	ld=sorted(ld,key=lambda _dx: -sum([abs(math.log10(v)) for v in _dx.values() if type(v)==float and v]))
	todel=['inf']
	for d in ld:
		if type(d['word'])==float:
			try:
				d['word']=str(int(d['word']))
			except (ValueError,OverflowError) as e:
				continue

		d['word']=escape_punc(d['word'])
		if not d['word']: continue
		if 'greek' in fn and ascii(d['word'])==d['word']: todel+=[d['word']]
		for k in d:
			if not d[k]: d[k]=1.0
			if type(d[k])==float:
				print k, d[k]
				d[k]=math.log10(d[k]) if d[k] else 1
				print k, d[k]
				print
				if d[k]>maxscore: d[k]=maxscore
	ld=[d for d in ld if not d['word'] in todel]

	fn=fn.replace('fmt_wide.','').replace('fmt_long.','')
	for d in ld:
		for k in d.keys():
			if not k: del d[k]

	fn=fn.replace('.txt','.{0}.minobs={1}.txt'.format(cellcol.replace('/',''), minobs))
	r=rpyd2.RpyD2(ld,rownamecol='word',allcols=True,zero=1)
	r.pca(fn=fn.replace('.txt','.words.pca.pdf'))

	if classes:
		#r.t().kclust(fn=fn.replace('.txt','.classes.kclust.pdf'),k=2)
		r.pvclust(fn=fn.replace('.txt','.classes.pvclust.pdf'))

		#ld2=ld[:len(ld[0])-1]
		#r2=rpyd2.RpyD2(ld2,rownamecol='word',allcols=True,zero=1)
		#r2.t().pca(fn.replace('.txt','.classes.pca.pdf'))



##############################################################

def ld2bayes(ld,textkey='text',metakeys=[],samplekey=None,windowsize=100,ngrams=[1,2],save=True):
	## really, an ld2featsets
	feat_sets={}
	ld2=[]
	def addslice(feat,sliceid):
		if not feat in feat_sets: feat_sets[feat]=[]
		feat_sets[feat]+=[sliceid]
	
	for d in ld:
		if not textkey in d: continue
		print d
		slicenum=0
		words=tokenize2(d[textkey]) if type(d[textkey])!=list else d[textkey]
		for txt in slice(words,slice_length=windowsize,runts=False):
			slicenum+=1
			sliceid=(d['filename'], slicenum)
			
			for mkey in metakeys:
				mkeyval=unicode(mkey)+u':'+unicode(d[mkey])
				addslice(mkeyval,sliceid)
			
			## words
			for ngramlen in ngrams:
				for gram in ngram(txt,ngramlen):
					gramstr=u"_".join(gram)
					feat=u'{0}-gram:{1}'.format(ngramlen,gramstr)
					addslice(feat,sliceid)
	
	for feat in feat_sets:
		feat_sets[feat]=set(feat_sets[feat])
	
	if save:
		import cPickle
		cPickle.dump(feat_sets,open('feat_sets.{0}.pickle'.format('_'.join(sorted(ld[0].keys()))[:25]),'wb'))
	else:
		return sets2bayes(feat_sets)

def sets2bayes(feat_sets, ofn=None,filter_by=None, minobs=10, stopwords=['1-gram:p','2-gram:p_m','1-gram:m']):
	import os,cPickle,pystats,networkx as nx,random
	for sw in stopwords:
		if not sw in feat_sets: continue
		del feat_sets[sw]
	
	def cmp(k1,k2,feat_sets,allnums=None):
		numposs=len(allnums)
		## test whether k1's being the case increases the odds of k2's occurring
		ids_k1_occurring=set(feat_sets[k1])
		#allnums=set(range(numposs))
		ids_k1_NOToccurring=allnums-ids_k1_occurring

		minnum=min([len(ids_k1_occurring), len(ids_k1_NOToccurring)])
		if not minnum or minnum<30: return
		#control_NOTk1=set(random.sample(list(ids_k1_NOToccurring),minnum))
		#study_k1=set(random.sample(list(ids_k1_occurring),minnum))	

		study_k1=ids_k1_occurring
		control_NOTk1=ids_k1_NOToccurring

		ids_k2_occurring=set(feat_sets[k2])
		k2_given_k1=len(study_k1 & ids_k2_occurring)
		notk2_given_k1=minnum-k2_given_k1
		p_k2_given_k1=k2_given_k1/len(study_k1)
		k2_given_notk1=len(control_NOTk1 & ids_k2_occurring)
		notk2_given_notk1=minnum-k2_given_notk1
		p_k2_given_notk1=k2_given_notk1/len(control_NOTk1)
		dx={}
		dx['Y_name']=k2
		dx['X_name']=k1
		dx['odds_Y_given_X']='{0} out of {1}'.format(k2_given_k1,len(study_k1))
		dx['odds_Y_given_notX']='{0} out of {1}'.format(k2_given_notk1,len(control_NOTk1))
		dx['odds_ratio2']=p_k2_given_k1/p_k2_given_notk1 if p_k2_given_notk1 else 0
		dx['odds_diff']=p_k2_given_k1 - len(ids_k2_occurring)/numposs
		dx['prob_X']=len(ids_k1_occurring)/numposs
		dx['prob_Y']=len(ids_k2_occurring)/numposs
		dx['prob_Y_given_X']=p_k2_given_k1
		dx['prob_Y_given_notX']=p_k2_given_notk1
		dx['num_X']=len(feat_sets[k1])
		dx['num_Y']=len(feat_sets[k2])
		dx['num_X_and_Y']=len(feat_sets[k1] & feat_sets[k2])
		if not dx['prob_Y_given_X'] or not dx['prob_Y_given_notX']: return
		return dx

	
	def do_featset(setfn, feat_sets):
		G=nx.DiGraph()
		print len(feat_sets)
		allnums=set(item for sublist in feat_sets.values() for item in sublist)
		numposs=len(allnums)
		i=0
		ii=0
		ilim=len(feat_sets)
		for k1,k2 in sorted(pystats.product(feat_sets.keys(),feat_sets.keys())):
			print k1,k2,len(feat_sets[k1]),len(feat_sets[k2]),'...'
			if k1==k2: continue
			k1x=set(k1.split(':',1)[1].split('_'))
			k2x=set(k2.split(':',1)[1].split('_'))
			k1prefix=k1.split(':')[0]
			k2prefix=k2.split(':')[0]
			#if k1prefix == k2prefix:
			#	continue

			if {k1prefix,k2prefix} in [ {'Type','Era'}, {'Era','Booth'}, {'Type','Booth'}, {'Space','Era'}, {'Space','Type'} ]:
				continue

			if len(k1x&k2x):
				print ">> skipping:",k1,k2
				continue
			
			
			d=cmp(k1,k2,feat_sets,allnums=allnums)
			if not d: continue
			#print d,'...'
			print ilim,ii,i
			print
			i+=1
			if i>ilim:
				i=0
				ii+=1
			if d['odds_diff']<0: continue
			for kx in [k1,k2]:
				_a,_b=kx.split(':',1)
				if not G.has_node(kx): G.add_node(kx,nodeType=_a,nodeName=_b,nodeFreq=len(feat_sets[kx]))
			G.node[k1]['nodeProb']=d['prob_X']
			G.node[k2]['nodeProb']=d['prob_Y']
			print G.node[k1]
			G.add_edge(k1,k2,weight=d['odds_ratio2'],Label=d['odds_ratio2'],**d)
		nx.write_gexf(G, setfn+'.gexf')
		
		return G


	if type(feat_sets) in [str,unicode]:
		print ">> loading sets:"
		feat_sets=cPickle.load(open(feat_sets))
		print ">> done."
	
	
	
	if not ofn: ofn='bayesnet.'+now()
	
	if filter_by:
		relevkeys=[k for k in feat_sets.keys() if k.startswith(filter_by)]
		minlen=min([len(feat_sets[key]) for key in relevkeys])
		for key in relevkeys:
			featsetnow={}
			allowablekeys=set(random.sample(list(feat_sets[key]), minlen))
			for kx in feat_sets.keys():
				setnow=set(list(feat_sets[kx]))&allowablekeys
				#if len(setnow)<lenlim/2: continue
				if len(setnow)<minobs: continue
				print kx, len(setnow)
				featsetnow[kx]=setnow
			print len(featsetnow)
			return do_featset(ofn+'.'+key, featsetnow)
	else:
		for k in feat_sets.keys():
			if len(feat_sets[k])<minobs:
				del feat_sets[k]

		print ">> # keys to analyze:", len(feat_sets)
		print feat_sets.keys()
		
		return do_featset(ofn, feat_sets)			

	
##############################################################

def binyear(year,binsize,start=0):
	if binsize==1: return '{0}'.format(year)
	for a,b in bigrams(range(start,3000,binsize)):
		if year in range(a,b):
			return '{0}-{1}'.format(a,b-1)
	return '????-????'

def bin(num,step,start=0.0,max=None,zfill=3):
	numnow=start
	while numnow + step < num:
		if max and numnow + step>max: break
		numnow+=step
		

	if int(numnow) == numnow: numnow=int(numnow)
	if max and numnow + step >max: return str(numnow).zfill(zfill)+'-'
	return str(numnow).zfill(zfill)+'-'+str(numnow+step-1)




def toks2freq(toks,tfy=False):
	tokd={}
	for tok in toks:
		try:
			tokd[tok]+=1
		except:
			tokd[tok]=1

	if tfy:
		import pystats
		return pystats.tfy(tokd)
	else:
		return tokd

def text2toks(text):
	toks=[]
	for w in text.lower().split():
		for ww in w.split('-'):
			www=noPunc(ww)
			toks.append(www)
	return toks

def hash(string):
	import hashlib
	return str(hashlib.sha224(string).hexdigest())


def tokenize(speech,regex=r'(\s+|\w+|\S+|\W+)'):
	c=re.compile(regex,re.U)
	l=[]
	for w in [noPunc(x).lower() for x in c.findall(speech) if noPunc(x).strip()]:
		for ww in w.split('.'):
			if ww: l+=[ww]
	return l

def tokenize2(speech,punc=False,regex=r'(\s+|\w+|\S+|\W+)',lower=True):
	c=re.compile(regex,re.U)
	l=[]
	for w in [(x.lower() if lower else x) for x in c.findall(speech) if x.strip()]:
		wx=''
		if (w.startswith("'") or w.startswith("-")) and len(l):
			l[-1]+=w
			continue
		
		for ww in w:
			if ww.isalpha() or (wx and ww in [u"'",u"-"]):
				wx+=ww
			else:
				if wx: l+=[wx]
				if punc: l+=[ww]
				wx=''
		if wx: l+=[wx]
	return l

def sortd(dict):
	return sorted(dict.items(), key=lambda lx: -lx[1])



def crunch(objects,function_or_methodname,ismethod=None,nprocs=16,args=[],kwargs={}):
	import time,random
	ismethod=type(function_or_methodname) in [str,unicode] if ismethod is None else ismethod 
	def do_preparse(text,args=[],kwargs={}):
		threadid=os.getpid()
		time.sleep(random.uniform(0,5))
		print "[{2}] Starting working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid)
		print ismethod,function_or_methodname,args,kwargs
		if ismethod:
			x=getattr(text,function_or_methodname)(*args,**kwargs)
		else:
			x=function_or_methodname(text, *args, **kwargs)

		print "[{2}] Finished working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid)
		return x

	import thread,multiprocessing,os
	from multiprocessing import Process, Pipe
	from itertools import izip

	def spawn(f):
		def fun(q_in,q_out):
			numdone=0
			while True:
				numdone+=1
				i,x = q_in.get()
				if i == None:
					break
				q_out.put((i,f(x,args=args,kwargs=kwargs)))
		return fun

	def parmap(f, X, nprocs = multiprocessing.cpu_count()):
		q_in   = multiprocessing.Queue(1)
		q_out  = multiprocessing.Queue()

		proc = [multiprocessing.Process(target=spawn(f),args=(q_in,q_out)) for _ in range(nprocs)]
		for p in proc:
			p.daemon = True
			p.start()

		sent = [q_in.put((i,x)) for i,x in enumerate(X)]
		[q_in.put((None,None)) for _ in range(nprocs)]
		res = [q_out.get() for _ in range(len(sent))]

		[p.join() for p in proc]

		return [x for i,x in sorted(res)]

	parmap(do_preparse, objects, nprocs=nprocs)
	return True




import re

#Define exceptions
class RomanError(Exception): pass
class OutOfRangeError(RomanError): pass
class NotIntegerError(RomanError): pass
class InvalidRomanNumeralError(RomanError): pass

#Define digit mapping
romanNumeralMap = (('M',  1000),
				   ('CM', 900),
				   ('D',  500),
				   ('CD', 400),
				   ('C',  100),
				   ('XC', 90),
				   ('L',  50),
				   ('XL', 40),
				   ('X',  10),
				   ('IX', 9),
				   ('V',  5),
				   ('IV', 4),
				   ('I',  1))

def toRoman(n):
	"""convert integer to Roman numeral"""
	if not (0 < n < 5000):
		raise OutOfRangeError, "number out of range (must be 1..4999)"
	if int(n) != n:
		raise NotIntegerError, "decimals can not be converted"

	result = ""
	for numeral, integer in romanNumeralMap:
		while n >= integer:
			result += numeral
			n -= integer
	return result

#Define pattern to detect valid Roman numerals
romanNumeralPattern = re.compile("""
	^					# beginning of string
	M{0,4}				# thousands - 0 to 4 M's
	(CM|CD|D?C{0,3})	# hundreds - 900 (CM), 400 (CD), 0-300 (0 to 3 C's),
						#			 or 500-800 (D, followed by 0 to 3 C's)
	(XC|XL|L?X{0,3})	# tens - 90 (XC), 40 (XL), 0-30 (0 to 3 X's),
						#		 or 50-80 (L, followed by 0 to 3 X's)
	(IX|IV|V?I{0,3})	# ones - 9 (IX), 4 (IV), 0-3 (0 to 3 I's),
						#		 or 5-8 (V, followed by 0 to 3 I's)
	$					# end of string
	""" ,re.VERBOSE)

def fromRoman(s):
	"""convert Roman numeral to integer"""
	if not s:
		raise InvalidRomanNumeralError, 'Input can not be blank'
	if not romanNumeralPattern.search(s):
		raise InvalidRomanNumeralError, 'Invalid Roman numeral: %s' % s

	result = 0
	index = 0
	for numeral, integer in romanNumeralMap:
		while s[index:index+len(numeral)] == numeral:
			result += integer
			index += len(numeral)
	return result



#!/usr/bin/env python
#coding:utf-8
# Author: Alejandro Nolla - z0mbiehunt3r
# Purpose: Example for detecting language using a stopwords based approach
# Created: 15/05/13



#----------------------------------------------------------------------
def _calculate_languages_ratios(text):
	"""
	Calculate probability of given text to be written in several languages and
	return a dictionary that looks like {'french': 2, 'spanish': 4, 'english': 0}
	
	@param text: Text whose language want to be detected
	@type text: str
	
	@return: Dictionary with languages and unique stopwords seen in analyzed text
	@rtype: dict
	"""

	from nltk import wordpunct_tokenize
	from nltk.corpus import stopwords


	languages_ratios = {}

	'''
	nltk.wordpunct_tokenize() splits all punctuations into separate tokens
	
	>>> wordpunct_tokenize("That's thirty minutes away. I'll be there in ten.")
	['That', "'", 's', 'thirty', 'minutes', 'away', '.', 'I', "'", 'll', 'be', 'there', 'in', 'ten', '.']
	'''

	tokens = wordpunct_tokenize(text)
	words = [word.lower() for word in tokens]

	# Compute per language included in nltk number of unique stopwords appearing in analyzed text
	for language in stopwords.fileids():
		stopwords_set = set(stopwords.words(language))
		words_set = set(words)
		common_elements = words_set.intersection(stopwords_set)

		languages_ratios[language] = len(common_elements) # language "score"

	return languages_ratios


#----------------------------------------------------------------------
def detect_language(text):
	"""
	Calculate probability of given text to be written in several languages and
	return the highest scored.
	
	It uses a stopwords based approach, counting how many unique stopwords
	are seen in analyzed text.
	
	@param text: Text whose language want to be detected
	@type text: str
	
	@return: Most scored language guessed
	@rtype: str
	"""

	import sys
	ratios = _calculate_languages_ratios(text)

	most_rated_language = max(ratios, key=ratios.get)

	return most_rated_language







def morphadorn(text):

	"""text=Mary+had+a+lytle+lamb+whose+vertue+was+strykinge+as+snowe&
	corpusConfig=eme&
	includeInputText=true&
	media=json&
	xmlOutputType=outputPlainXML&
	adorn=Adorn"""

	from simplejson import JSONDecodeError
	import requests
	rd={}
	rd['corpusConfig']='eme'
	rd['text']=text
	rd['media']='json'
	rd['adorn']='Adorn'
	rd['xmlOutputType']='outputPlainXML'
	rd['includeInputText']=False
	import random
	from requests import ConnectionError
	while True:
		port=random.choice(['8182','8183'])
		print ">> morphadorning on port",port
		try:
			r=requests.post('http://localhost:'+port+'/partofspeechtagger',data=rd)
			return r.json()
		except (JSONDecodeError,ConnectionError) as e:
			print "!!",e
			print "!! failed to get data from morphadorn server, trying again after a short nap"
			from random import randint
			from time import sleep
			naptime=randint(1,60)
			print ">> napping for {0}sec...".format(naptime)
			sleep(naptime)
			print ">> awake."


def mongosizes():
	import pymongo
	db=pymongo.Connection().litlab
	for cname in db.collection_names():
		cdb=getattr(db,cname)




color_hexes=['#4B0082', '#0000CD', '#228B22', '#C71585', '#FF00FF', '#9400D3', '#20B2AA', '#90EE90', '#32CD32', '#FF6347', '#9370DB', '#A0522D', '#8A2BE2', '#191970', '#6A5ACD', '#808080', '#483D8B', '#5F9EA0', '#DC143C', '#DB7093', '#FF0000', '#3CB371', '#8B008B', '#EE82EE', '#FFB6C1', '#008080', '#48D1CC', '#AFEEEE', '#8FBC8F', '#00FF00', '#708090', '#00FFFF', '#9ACD32', '#7FFFD4', '#8B4513', '#7B68EE', '#FF00FF', '#00FA9A', '#FF7F50', '#9932CC', '#DEB887', '#CD853F', '#BDB76B', '#0000FF', '#6B8E23', '#2F4F4F', '#66CDAA', '#FF8C00', '#B8860B', '#FFC0CB', '#FA8072', '#ADD8E6', '#FFA500', '#696969', '#87CEEB', '#40E0D0', '#B0C4DE', '#CD5C5C', '#F08080', '#4682B4', '#FF69B4', '#00FFFF', '#8B0000', '#800000', '#BC8F8F', '#7FFF00', '#00FF7F', '#00BFFF', '#7CFC00', '#006400', '#008000', '#FFDAB9', '#D2691E', '#008B8B', '#4169E1', '#556B2F', '#B22222', '#FFDEAD', '#C0C0C0', '#F0E68C', '#F5DEB3', '#E9967A', '#DA70D6', '#1E90FF', '#ADFF2F', '#00008B', '#B0E0E6', '#FAF0E6', '#FFE4B5', '#FFFAFA', '#000000', '#F5F5F5', '#E6E6FA', '#FFEFD5', '#FFA07A', '#00CED1', '#FFFFE0', '#BA55D3', '#F5F5DC', '#FFFF00', '#DAA520', '#808000', '#000080', '#A9A9A9', '#F4A460', '#FF1493', '#A52A2A', '#D8BFD8', '#DDA0DD', '#800080', '#98FB98', '#2E8B57', '#D2B48C', '#87CEFA', '#FF4500', '#6495ED', '#FAFAD2', '#778899']